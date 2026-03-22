"""
Export optimizer results to a color-coded Excel workbook.

Each press location gets two tabs:
  - <location>         : combined jobs (2+ per combo), color-coded by ComboID
  - <location>_SINGLES : single jobs (no combination possible)
"""

import sys
import subprocess
import pandas as pd

PALETTE = [
    "#FFF2CC", "#D9E1F2", "#FCE4D6", "#E2EFDA", "#EDEDED",
    "#E6E0EC", "#F8CBAD", "#C6E0B4", "#CFE2F3", "#FFD966",
    "#DDEBF7", "#EAD1DC", "#D0E0E3", "#F4CCCC", "#CCE5FF",
]


def _sanitize_sheet(name: str) -> str:
    s = str(name or "Unknown").strip() or "Unknown"
    for ch in r'[]:*?/\\':
        s = s.replace(ch, "_")
    return s[:31]


def _hex_to_argb(h: str) -> str:
    h = (h or "#FFFFFF").lstrip("#").upper()
    return "FF" + (h if len(h) == 6 else "FFFFFF")


def _as_int(v) -> int:
    try:
        return int(float(str(v).replace(",", "").strip()))
    except Exception:
        return 0


def export_to_excel(total_runs_by_loc: dict, jobs: list, path: str = "combo_output.xlsx"):
    """
    Export optimizer results to a colored Excel workbook.

    Args:
        total_runs_by_loc: {press_location: {combo_id: [job_ids]}}
        jobs: list of PrintJob objects
        path: output file path
    """
    try:
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter

    job_lookup = {
        j.JOB: {
            "PRODUCTTYPE": j.PRODUCTTYPE,
            "MANUFACTURING_LOCATION": j.PRESS_LOCATION,
            "SEND_TO_LOCATION": j.SEND_TO_LOCATION,
            "PAPER": j.PAPER,
            "FINISHTYPE": j.FINISHTYPE,
            "FINISHINGOP": j.FINISHINGOP,
            "SHIPPED_DATE": j.DELIVERYDATE,
            "INKSS1": j.INKSS1,
            "INKSS2": j.INKSS2,
            "QUANTITYORDERED": _as_int(j.QUANTITYORDERED),
            "PAGES": _as_int(j.PAGES),
        }
        for j in jobs
    }

    used_names = {}

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for loc, runs_loc in total_runs_by_loc.items():
            rows = []
            for combo_id, job_ids in runs_loc.items():
                for job_id in job_ids:
                    m = job_lookup.get(job_id, {})
                    rows.append({"ComboID": combo_id, "JOB": job_id, **m})

            if not rows:
                continue

            df_x = pd.DataFrame(rows).sort_values(["ComboID", "JOB"]).reset_index(drop=True)
            counts = df_x.groupby("ComboID")["JOB"].transform("count")
            df_combos = df_x[counts > 1].copy()
            df_singles = df_x[counts == 1].copy()

            if not df_combos.empty:
                combo_map = {old: i + 1 for i, old in enumerate(sorted(df_combos["ComboID"].unique()))}
                df_combos["ComboID"] = df_combos["ComboID"].map(combo_map)

            if not df_singles.empty:
                single_map = {old: i + 1 for i, old in enumerate(sorted(df_singles["ComboID"].unique()))}
                df_singles["ComboID"] = df_singles["ComboID"].map(single_map)

            base = _sanitize_sheet(loc)
            cnt = used_names.get(base, 0)
            used_names[base] = cnt + 1
            sheet_name = base if cnt == 0 else f"{base[:31 - len(str(cnt)) - 1]}_{cnt}"

            if not df_combos.empty:
                df_combos.to_excel(writer, index=False, sheet_name=sheet_name)
                ws = writer.sheets[sheet_name]
                ws.freeze_panes = "A2"

                headers = list(df_combos.columns)
                combo_col = headers.index("ComboID") + 1
                unique_combos = sorted(df_combos["ComboID"].unique())
                color_map = {cid: PALETTE[i % len(PALETTE)] for i, cid in enumerate(unique_combos)}

                for r in range(2, len(df_combos) + 2):
                    cid_val = ws.cell(row=r, column=combo_col).value
                    hex_c = color_map.get(cid_val, "#FFFFFF")
                    fill = PatternFill(fill_type="solid",
                                      start_color=_hex_to_argb(hex_c),
                                      end_color=_hex_to_argb(hex_c))
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=r, column=c).fill = fill

                for col_idx, col in enumerate(headers, 1):
                    col_letter = get_column_letter(col_idx)
                    max_len = max(len(str(col)), *(len(str(v)) for v in df_combos[col].fillna("").astype(str)))
                    ws.column_dimensions[col_letter].width = min(80, max(10, max_len + 2))

            if not df_singles.empty:
                singles_name = _sanitize_sheet(f"{loc[:23]}_SINGLES")
                cnt_s = used_names.get(singles_name, 0)
                used_names[singles_name] = cnt_s + 1
                sheet_singles = singles_name if cnt_s == 0 else f"{singles_name[:28]}_{cnt_s}"

                df_singles.to_excel(writer, index=False, sheet_name=sheet_singles)
                ws2 = writer.sheets[sheet_singles]
                ws2.freeze_panes = "A2"

                headers2 = list(df_singles.columns)
                for col_idx, col in enumerate(headers2, 1):
                    col_letter = get_column_letter(col_idx)
                    max_len = max(len(str(col)), *(len(str(v)) for v in df_singles[col].fillna("").astype(str)))
                    ws2.column_dimensions[col_letter].width = min(80, max(10, max_len + 2))

    print(f"Exported results to {path}")
